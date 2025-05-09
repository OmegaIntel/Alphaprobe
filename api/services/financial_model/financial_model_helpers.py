import csv
from typing import Dict, Any, List, Tuple, Optional
from io import BytesIO, StringIO
from openpyxl import load_workbook

# Classification patterns
SHEET_TYPE_PATTERNS = {
    "income_statement": [
        "income", "profit", "loss", "p&l", "p & l", "revenue", "earnings",
        "operating result", "statement of operations"
    ],
    "balance_sheet": [
        "balance", "financial position", "assets", "liabilities", "equity",
        "bs "
    ],
    "cash_flow": [
        "cash flow", "cashflow", "cash", "cf ", "statement of cash",
        "sources and uses"
    ],
    "assumptions": [
        "assumption", "input", "driver", "lever", "parameter", "config"
    ],
    "detail": [
        "detail", "breakdown", "analysis", "deep dive", "unit", "economics",
        "traffic", "sales detail", "user", "customer"
    ],
    "projection": [
        "projection", "forecast", "budget", "plan", "target"
    ]
}

# Standard financial metrics by sheet type
STANDARD_METRICS = {
    "income_statement": [
        "Revenue", "Cost of Goods Sold", "Gross Profit", "Operating Expenses", 
        "EBITDA", "Depreciation", "EBIT", "Interest Expense", 
        "Profit Before Tax", "Tax", "Net Income"
    ],
    "balance_sheet": [
        "Current Assets", "Cash", "Accounts Receivable", "Inventory", 
        "Fixed Assets", "Total Assets", "Current Liabilities", 
        "Accounts Payable", "Long Term Debt", "Total Liabilities", "Equity"
    ],
    "cash_flow": [
        "Operating Cash Flow", "Investing Cash Flow", "Financing Cash Flow",
        "Net Cash Flow", "Beginning Cash Balance", "Ending Cash Balance"
    ],
}

# Mapping of common alternative names for key metrics
METRIC_ALTERNATIVES = {
    "Revenue": ["sales", "turnover", "income", "top line", "gross revenue"],
    "EBITDA": ["operating profit", "operating income", "ebitda"],
    "Net Income": ["net profit", "net earnings", "bottom line", "pat", "profit after tax"],
    "Cash": ["cash and cash equivalents", "cash & cash equivalents", "cash in bank"],
    "Total Assets": ["assets", "total assets"],
    "Total Liabilities": ["liabilities", "total liabilities"],
    "Operating Cash Flow": ["cash from operations", "operating activities", "ocf"],
    "Cost of Goods Sold": ["cogs", "cost of sales", "cost of revenue", "direct costs"],
    "Gross Profit": ["gross margin"],
    "Operating Expenses": ["opex", "sg&a", "operating costs"]
}

def process_file(file_content: bytes, file_ext: str) -> Dict[str, Any]:
    """
    Process a file based on its extension and extract its structure.
    
    Args:
        file_content: The binary content of the file
        file_ext: The file extension (e.g., ".xlsx", ".csv")
        
    Returns:
        Dictionary containing the file structure information
    """
    if file_ext == ".csv":
        return process_csv_file(file_content)
    else:  # Excel file
        return process_excel_file(file_content)

def process_csv_file(file_content: bytes) -> Dict[str, Any]:
    """
    Process a CSV file and extract its structure.
    
    Args:
        file_content: The binary content of the CSV file
        
    Returns:
        Dictionary containing the CSV structure information
    """
    try:
        content_str = file_content.decode('utf-8')
        csv_reader = csv.reader(StringIO(content_str))
        
        # Get headers (first row)
        headers = next(csv_reader, [])
        
        # Get sample values (second row)
        sample_values = next(csv_reader, [])
        
        # If sample values row has fewer items than headers, pad with None
        if len(sample_values) < len(headers):
            sample_values.extend([None] * (len(headers) - len(sample_values)))
        
        # Count all rows
        row_count = 2  # Already read 2 rows
        for row in csv_reader:
            row_count += 1
        
        # Determine file type based on headers
        file_type = "unknown"
        for sheet_type, patterns in SHEET_TYPE_PATTERNS.items():
            headers_text = " ".join(str(h).lower() for h in headers if h)
            if any(pattern in headers_text for pattern in patterns):
                file_type = sheet_type
                break
        
        # Create a structure similar to what we'd get from Excel
        return {
            "workbook_properties": {
                "sheet_count": 1,
                "file_type": "csv"
            },
            "sheets": {
                "CSV Data": {
                    "sheet_properties": {
                        "index": 0,
                        "type": file_type,
                        "dimensions": f"A1:{chr(65 + len(headers) - 1)}{row_count}"
                    },
                    "headers": headers,
                    "sample_values": sample_values,
                    "formulas": {}  # CSV files don't contain formulas
                }
            }
        }
    except Exception as e:
        print(f"Error processing CSV file: {str(e)}")
        # Return a minimal structure
        return {
            "workbook_properties": {
                "sheet_count": 1,
                "file_type": "csv",
                "error": str(e)
            },
            "sheets": {
                "CSV Data": {
                    "sheet_properties": {
                        "index": 0,
                        "type": "unknown"
                    },
                    "headers": [],
                    "sample_values": []
                }
            }
        }

def process_excel_file(file_content: bytes) -> Dict[str, Any]:
    """
    Process an Excel file and extract its structure.
    
    Args:
        file_content: The binary content of the Excel file
        
    Returns:
        Dictionary containing the Excel workbook structure information
    """
    try:
        # First load with data_only=True to get calculated values
        wb_input_values = load_workbook(BytesIO(file_content), data_only=True)
        
        # Then try to load again with data_only=False to access formulas
        try:
            wb_input_formulas = load_workbook(BytesIO(file_content), data_only=False)
            print("✅ Successfully loaded workbook with formulas")
            
            # Extract workbook structure including sheets, headers, and formulas
            workbook_data = extract_workbook_structure(wb_input_formulas)
            
            # Add sample values from the data_only=True workbook
            for sheet_name in workbook_data["sheets"]:
                if sheet_name in wb_input_values.sheetnames:
                    try:
                        data_sheet = wb_input_values[sheet_name]
                        # Get sample values from second row
                        sample_values = []
                        second_row = list(data_sheet.iter_rows(min_row=2, max_row=2, values_only=True))
                        if second_row and second_row[0]:
                            sample_values = list(second_row[0])
                            
                        workbook_data["sheets"][sheet_name]["sample_values"] = sample_values
                    except Exception as e:
                        print(f"Warning: Could not get values from sheet '{sheet_name}': {str(e)}")
                        
        except Exception as e:
            print(f"⚠️ Could not load workbook with formulas: {str(e)}")
            print("⚠️ Proceeding with calculated values only")
            # Fall back to just using the data_only=True workbook
            workbook_data = extract_workbook_structure(wb_input_values)
            
        return workbook_data
    except Exception as e:
        print(f"Error processing Excel file: {str(e)}")
        # Return a minimal structure
        return {
            "workbook_properties": {
                "sheet_count": 0,
                "file_type": "excel",
                "error": str(e)
            },
            "sheets": {}
        }

def classify_sheet_type(sheet_name: str, headers: List[str]) -> str:
    """
    Determine the type of financial sheet based on name and headers.
    
    Args:
        sheet_name: The name of the sheet
        headers: List of column headers in the sheet
        
    Returns:
        Classified sheet type (e.g., "income_statement", "balance_sheet")
    """
    sheet_name_lower = str(sheet_name).lower()
    
    # First check sheet name against patterns
    for sheet_type, patterns in SHEET_TYPE_PATTERNS.items():
        if any(pattern in sheet_name_lower for pattern in patterns):
            return sheet_type
    
    # If not found by name, check headers
    headers_text = " ".join(str(h).lower() for h in headers if h)
    for sheet_type, patterns in SHEET_TYPE_PATTERNS.items():
        if any(pattern in headers_text for pattern in patterns):
            return sheet_type
    
    # Default to unknown if no match found
    return "unknown"

def find_cell_formula(sheet, cell_reference: str) -> Optional[str]:
    """
    Safely extract formula from a cell if it exists.
    
    Args:
        sheet: The worksheet object
        cell_reference: Cell reference (e.g., "A1")
        
    Returns:
        The formula as a string, or None if no formula exists
    """
    try:
        cell = sheet[cell_reference]
        return cell.formula if hasattr(cell, 'formula') and cell.formula else None
    except:
        return None

def extract_workbook_structure(workbook) -> Dict[str, Any]:
    """
    Extract the structure of an Excel workbook including sheets, headers, and formulas.
    
    Args:
        workbook: The openpyxl workbook object
        
    Returns:
        Dictionary containing the workbook structure information
    """
    result = {
        "workbook_properties": {
            "sheet_count": len(workbook.sheetnames),
            "sheet_names": workbook.sheetnames,
            "file_type": "excel"
        },
        "sheets": {}
    }
    
    # Process each sheet in the workbook
    for idx, sheet_name in enumerate(workbook.sheetnames):
        try:
            sheet = workbook[sheet_name]
            
            # Get sheet dimensions
            dimensions = sheet.dimensions
            
            # Extract headers (first row)
            headers = []
            for cell in next(sheet.rows):
                headers.append(cell.value)
            
            # Extract sample values (second row for context)
            sample_values = []
            second_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
            if second_row:
                sample_values = list(second_row[0])
            
            # Determine sheet type based on name and headers
            sheet_type = classify_sheet_type(sheet_name, headers)
            
            # Extract formulas (sample from the sheet)
            # Note: This only works if workbook was loaded with data_only=False
            formulas = {}
            formula_count = 0
            
            # We'll try to get formulas, but won't fail if they're not available
            try:
                max_formulas = 20  # Limit number of formulas to extract
                
                # Define a max range to scan for formulas
                max_row = min(sheet.max_row, 50)  # Limit to first 50 rows for performance
                max_col = min(sheet.max_column, 20)  # Limit to first 20 columns
                
                for row in range(1, max_row + 1):
                    for col in range(1, max_col + 1):
                        cell = sheet.cell(row=row, column=col)
                        if hasattr(cell, 'formula') and cell.formula:
                            cell_ref = cell.coordinate
                            formulas[cell_ref] = cell.formula
                            formula_count += 1
                            
                            if formula_count >= max_formulas:
                                break
                    if formula_count >= max_formulas:
                        break
            except Exception as formula_error:
                print(f"Note: Could not extract formulas from sheet '{sheet_name}': {str(formula_error)}")
                # Continue processing without formulas
            
            # Store sheet data
            result["sheets"][sheet_name] = {
                "sheet_properties": {
                    "index": idx,
                    "type": sheet_type,
                    "dimensions": dimensions
                },
                "headers": headers,
                "sample_values": sample_values,
                "formulas": formulas
            }
            
            # Look for key cells with important data (like totals or key metrics)
            key_cells = {}
            
            # Try to identify key financial metrics based on row headers
            try:
                for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(30, sheet.max_row)), 1):
                    cell_value = row[0].value
                    if cell_value and isinstance(cell_value, str):
                        cell_value_lower = cell_value.lower()
                        
                        # Check for key financial terms
                        for metric, alternatives in METRIC_ALTERNATIVES.items():
                            if (metric.lower() in cell_value_lower or 
                                any(alt in cell_value_lower for alt in alternatives)):
                                
                                # Found a key metric row, record the cells in this row
                                for col_idx, cell in enumerate(row[1:], 1):
                                    key_cells[cell.coordinate] = {
                                        "metric": metric,
                                        "cell_value": cell.value
                                    }
            except Exception as e:
                print(f"Note: Error identifying key cells in '{sheet_name}': {str(e)}")
            
            result["sheets"][sheet_name]["key_cells"] = key_cells
            
        except Exception as e:
            print(f"Error extracting data from sheet '{sheet_name}': {str(e)}")
            # Include a minimal entry for this sheet
            result["sheets"][sheet_name] = {
                "sheet_properties": {
                    "index": idx,
                    "type": "error",
                    "error": str(e)
                }
            }
    
    return result

def find_standardized_cell_mapping(reference_model_mapping: Dict, update_file_mapping: Dict) -> Dict[str, Dict[str, Any]]:
    """
    Create a mapping between standardized metrics and cells in the template.
    
    Args:
        reference_model_mapping: The mapping JSON from the reference model
        update_file_mapping: The mapping JSON from the update file
        
    Returns:
        Dictionary mapping metrics to cell locations in the reference template
    """
    cell_mapping = {}
    
    try:
        # Extract metrics from the reference model
        for sheet_name, sheet_data in reference_model_mapping.get("sheets", {}).items():
            key_metrics = sheet_data.get("key_metrics", {})
            
            for metric_name, metric_data in key_metrics.items():
                cell_references = metric_data.get("found_in_cells", [])
                source_column = metric_data.get("source_column")
                
                if cell_references:
                    # We have specific cell references for this metric
                    cell_mapping[metric_name] = {
                        "sheet": sheet_name,
                        "cells": cell_references,
                        "source_column": source_column
                    }
                elif "key_cells" in sheet_data:
                    # Try to find cells from key_cells
                    for cell_ref, cell_info in sheet_data.get("key_cells", {}).items():
                        if cell_info.get("metric") == metric_name:
                            if metric_name not in cell_mapping:
                                cell_mapping[metric_name] = {
                                    "sheet": sheet_name,
                                    "cells": [],
                                    "source_column": source_column
                                }
                            cell_mapping[metric_name]["cells"].append(cell_ref)
        
        return cell_mapping
    except Exception as e:
        print(f"Error creating cell mapping: {str(e)}")
        return {}

def update_template_with_new_data(
    template_file_path: str, 
    cell_mapping: Dict[str, Dict[str, Any]], 
    update_file_mapping: Dict
) -> Tuple[bool, List[str]]:
    """
    Update the template workbook with new data from the update file.
    
    Args:
        template_file_path: Path to the template Excel file
        cell_mapping: Mapping between metrics and cells in the template
        update_file_mapping: The mapping JSON from the update file
        
    Returns:
        Tuple of (success_flag, list_of_updates_made)
    """
    updated_cells = []
    
    try:
        # Load the template workbook
        template_workbook = load_workbook(template_file_path)
        
        # Process each sheet in the update file
        for sheet_name, sheet_data in update_file_mapping.get("sheets", {}).items():
            metrics = sheet_data.get("metrics", {})
            
            # For each metric in the update file
            for metric_name, metric_data in metrics.items():
                # Check if we have a mapping for this metric
                if metric_name in cell_mapping:
                    # Get the mapping info
                    map_info = cell_mapping[metric_name]
                    template_sheet_name = map_info["sheet"]
                    cell_refs = map_info["cells"]
                    
                    # Get the value from the update file
                    new_value = metric_data.get("sample_value")
                    
                    # Only proceed if we have a valid value
                    if new_value is not None:
                        # Update the template
                        if template_sheet_name in template_workbook.sheetnames:
                            template_sheet = template_workbook[template_sheet_name]
                            
                            # Update all mapped cells for this metric
                            for cell_ref in cell_refs:
                                template_sheet[cell_ref] = new_value
                                updated_cells.append(f"{template_sheet_name}!{cell_ref} ({metric_name}): {new_value}")
        
        # Save the updated template
        template_workbook.save(template_file_path)
        
        return len(updated_cells) > 0, updated_cells
    except Exception as e:
        print(f"Error updating template: {str(e)}")
        return False, [f"Error: {str(e)}"]