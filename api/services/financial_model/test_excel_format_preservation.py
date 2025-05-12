import os
import sys
import pytest
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from pathlib import Path

# Add project root to sys.path
sys.path.append(str(Path(__file__).parent.parent))

# Import the Excel template preserver
from services.financial_model.excel_format_preservation import ExcelTemplatePreserver

def create_test_template():
    """Create a test Excel template with various formatting and formulas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Income Statement"
    
    # Create header row with formatting
    headers = ["Year", "Revenue", "COGS", "Gross Profit", "OpEx", "EBITDA", "Tax", "Net Income"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12, color="0000FF")
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"), 
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    # Add data for 2023
    data_2023 = [2023, 1000000, 400000, "=B2-C2", 200000, "=D2-E2", "=0.25*F2", "=F2-G2"]
    for col, value in enumerate(data_2023, 1):
        cell = ws.cell(row=2, column=col, value=value)
        if isinstance(value, str) and value.startswith("="):
            cell.font = Font(italic=True)
    
    # Add data for 2024 with different formatting
    data_2024 = [2024, 1200000, 450000, "=B3-C3", 220000, "=D3-E3", "=0.25*F3", "=F3-G3"]
    for col, value in enumerate(data_2024, 1):
        cell = ws.cell(row=3, column=col, value=value)
        if isinstance(value, str) and value.startswith("="):
            cell.font = Font(italic=True)
            
        # Make revenue column green text for growth
        if col == 2:  # Revenue column
            cell.font = Font(color="00FF00", bold=True)
    
    # Add a Balance Sheet tab
    ws2 = wb.create_sheet(title="Balance Sheet")
    bs_headers = ["Year", "Cash", "AR", "Inventory", "Total Assets", "AP", "Debt", "Total Liabilities", "Equity"]
    
    for col, header in enumerate(bs_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="EEEEFF", end_color="EEEEFF", fill_type="solid")
    
    # Add data to Balance Sheet
    bs_data = [2023, 500000, 200000, 150000, "=SUM(B2:D2)", 100000, 300000, "=SUM(F2:G2)", "=E2-H2"]
    for col, value in enumerate(bs_data, 1):
        cell = ws2.cell(row=2, column=col, value=value)
    
    # Add conditional formatting to the Income Statement sheet
    from openpyxl.formatting.rule import ColorScaleRule
    ws.conditional_formatting.add(
        'B2:B3',
        ColorScaleRule(
            start_type='min', start_color='FFFFFF',
            end_type='max', end_color='00FF00'
        )
    )
    
    # Set column widths
    for i, width in enumerate([10, 15, 15, 15, 15, 15, 15, 15], 1):
        ws.column_dimensions[chr(64 + i)].width = width
        if i < len(bs_headers) + 1:
            ws2.column_dimensions[chr(64 + i)].width = width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def create_test_update_file():
    """Create a test update file with new data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue Update"
    
    # Create headers
    headers = ["Year", "Revenue", "COGS", "OpEx", "Tax Rate"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add updated data
    updated_data = [2024, 1500000, 600000, 250000, 0.25]
    for col, value in enumerate(updated_data, 1):
        ws.cell(row=2, column=col, value=value)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def test_excel_format_preservation():
    """Test the Excel format preservation functionality."""
    # Create test files
    template_content = create_test_template()
    update_content = create_test_update_file()
    
    # Create a template preserver
    preserver = ExcelTemplatePreserver(template_content)
    
    # Test loading the template
    assert preserver.template_workbook is not None
    assert "Income Statement" in preserver.template_workbook.sheetnames
    assert "Balance Sheet" in preserver.template_workbook.sheetnames
    
    # Create a simple mapping
    cell_mapping = {
        "Revenue": {
            "sheet": "Income Statement",
            "cells": ["B3"]  # Update 2024 Revenue
        },
        "COGS": {
            "sheet": "Income Statement",
            "cells": ["C3"]  # Update 2024 COGS
        },
        "OpEx": {
            "sheet": "Income Statement",
            "cells": ["E3"]  # Update 2024 OpEx
        }
    }
    
    # Create update file mapping
    update_file_mapping = {
        "sheets": {
            "Revenue Update": {
                "metrics": {
                    "Revenue": {
                        "source_column": "Revenue",
                        "sample_value": 1500000
                    },
                    "COGS": {
                        "source_column": "COGS",
                        "sample_value": 600000
                    },
                    "OpEx": {
                        "source_column": "OpEx",
                        "sample_value": 250000
                    }
                }
            }
        }
    }
    
    # Update the template
    success, updates = preserver.update_cells_with_mapping(cell_mapping, update_file_mapping)
    
    # Verify update was successful
    assert success
    assert len(updates) > 0
    
    # Save the updated workbook
    updated_content = preserver.save_workbook()
    
    # Load the updated workbook to verify changes
    updated_wb = load_workbook(BytesIO(updated_content))
    updated_ws = updated_wb["Income Statement"]
    
    # Check that values were updated
    assert updated_ws["B3"].value == 1500000  # Revenue updated
    assert updated_ws["C3"].value == 600000   # COGS updated
    assert updated_ws["E3"].value == 250000   # OpEx updated
    
    # Check that formulas still exist and were not overwritten
    # Different versions of openpyxl have different ways to access formulas
    try:
        # First try the .formula attribute (newer versions)
        assert updated_ws["D3"].formula == "B3-C3"
        assert updated_ws["F3"].formula == "D3-E3"
        assert updated_ws["G3"].formula == "0.25*F3"
        assert updated_ws["H3"].formula == "F3-G3"
    except AttributeError:
        # Fall back to ._value which contains the formula in older versions
        # or try other attributes that might contain the formula
        d3_formula = getattr(updated_ws["D3"], "_value", None)
        if d3_formula is None:
            d3_formula = getattr(updated_ws["D3"], "value", "")
            # In some versions, formulas start with =
            if isinstance(d3_formula, str) and d3_formula.startswith("="):
                d3_formula = d3_formula[1:]  # Remove the equals sign
        
        # Check that the formula strings are present (might be in different formats)
        assert "B3-C3" in str(d3_formula)
        
        # For the other cells, just check that they contain calculations
        assert isinstance(updated_ws["D3"].value, (int, float))  # Should be a calculated value
        assert isinstance(updated_ws["F3"].value, (int, float))  # Should be a calculated value
        assert isinstance(updated_ws["G3"].value, (int, float))  # Should be a calculated value
        assert isinstance(updated_ws["H3"].value, (int, float))  # Should be a calculated value
    
    # Check that formatting was preserved
    assert updated_ws["B3"].font.color.rgb == "00FF00"  # Green color for revenue was preserved
    
    # Check that conditional formatting still exists
    assert len(updated_ws.conditional_formatting) > 0
    
    # Check that column widths were preserved
    assert updated_ws.column_dimensions["B"].width == 15
    
    print("✅ All tests passed: Excel format preservation is working correctly!")
    return True

def test_excel_template_preserver_static_method():
    """Test the static method for template processing."""
    # Create test files
    template_content = create_test_template()
    update_content = create_test_update_file()
    
    # Create a simple mapping
    cell_mapping = {
        "Revenue": {
            "sheet": "Income Statement",
            "cells": ["B3"]  # Update 2024 Revenue
        },
        "COGS": {
            "sheet": "Income Statement",
            "cells": ["C3"]  # Update 2024 COGS
        },
        "OpEx": {
            "sheet": "Income Statement",
            "cells": ["E3"]  # Update 2024 OpEx
        }
    }
    
    # Create update file mapping
    update_file_mapping = {
        "sheets": {
            "Revenue Update": {
                "metrics": {
                    "Revenue": {
                        "source_column": "Revenue",
                        "sample_value": 1500000
                    },
                    "COGS": {
                        "source_column": "COGS",
                        "sample_value": 600000
                    },
                    "OpEx": {
                        "source_column": "OpEx",
                        "sample_value": 250000
                    }
                }
            }
        }
    }
    
    # Use the static method
    updated_content, updates = ExcelTemplatePreserver.process_template_update(
        template_content, update_file_mapping, cell_mapping
    )
    
    # Verify updates were returned
    assert len(updates) > 0
    
    # Load the updated workbook to verify changes
    updated_wb = load_workbook(BytesIO(updated_content))
    updated_ws = updated_wb["Income Statement"]
    
    # Check that values were updated
    assert updated_ws["B3"].value == 1500000  # Revenue updated
    
    # Check that formulas still exist
    try:
        # Try the .formula attribute (newer versions)
        assert updated_ws["D3"].formula == "B3-C3"
    except AttributeError:
        # Fall back to checking the calculated value 
        # Since we updated B3 and C3, D3 should be B3-C3 = 1500000-600000 = 900000
        assert updated_ws["D3"].value == 900000 or updated_ws["D3"].value == "=B3-C3"
    
    print("✅ Static method test passed: ExcelTemplatePreserver.process_template_update works correctly!")
    return True

if __name__ == "__main__":
    # Run tests
    test_excel_format_preservation()
    test_excel_template_preserver_static_method()
    print("All tests completed successfully!")