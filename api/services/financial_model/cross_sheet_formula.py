import re
import logging
from typing import Dict, Any, List, Set, Tuple
from openpyxl import load_workbook
from io import BytesIO

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger("cross_sheet_formulas")

class CrossSheetFormulaAnalyzer:
    """Analyzes and maintains cross-sheet formula relationships in Excel workbooks."""
    
    def __init__(self, workbook_content: bytes):
        """
        Initialize with workbook content.
        
        Args:
            workbook_content: Binary content of the Excel workbook
        """
        self.workbook_content = workbook_content
        self.workbook = None
        self.formula_map = {}
        self.sheet_dependencies = {}
        self.cell_dependencies = {}
        self.load_workbook()
        
    def load_workbook(self):
        """Load the workbook and initialize analysis."""
        try:
            self.workbook = load_workbook(BytesIO(self.workbook_content), data_only=False)
            logger.info(f"Loaded workbook with {len(self.workbook.sheetnames)} sheets for cross-sheet analysis")
        except Exception as e:
            logger.error(f"Error loading workbook: {str(e)}")
            raise Exception(f"Failed to load workbook for formula analysis: {str(e)}")
    
    def analyze_all_cross_sheet_formulas(self):
        """
        Analyze all formulas in the workbook to identify cross-sheet references.
        
        Returns:
            Dictionary of sheet dependencies and cell dependencies
        """
        logger.info("Starting analysis of cross-sheet formulas")
        
        # Reset maps
        self.formula_map = {}
        self.sheet_dependencies = {}
        self.cell_dependencies = {}
        
        # Process each sheet
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            # Initialize sheet in dependencies map
            if sheet_name not in self.sheet_dependencies:
                self.sheet_dependencies[sheet_name] = {
                    "depends_on": set(),
                    "referenced_by": set()
                }
                
            # Scan cells for formulas
            max_row = min(sheet.max_row, 500)  # Limit to first 500 rows for performance
            max_col = min(sheet.max_column, 50)  # Limit to first 50 columns
            
            formula_count = 0
            cross_sheet_count = 0
            
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    try:
                        cell = sheet.cell(row=row, column=col)
                        
                        # Get formula from cell (handling different openpyxl versions)
                        formula = self._get_cell_formula(cell)
                        
                        if formula:
                            formula_count += 1
                            cell_ref = f"{sheet_name}!{cell.coordinate}"
                            self.formula_map[cell_ref] = formula
                            
                            # Look for cross-sheet references
                            cross_refs = self._extract_sheet_references(formula)
                            
                            if cross_refs:
                                cross_sheet_count += 1
                                
                                # Add to cell dependencies
                                self.cell_dependencies[cell_ref] = {
                                    "formula": formula,
                                    "references": cross_refs
                                }
                                
                                # Add to sheet dependencies
                                for ref_sheet in cross_refs:
                                    # Current sheet depends on referenced sheet
                                    self.sheet_dependencies[sheet_name]["depends_on"].add(ref_sheet)
                                    
                                    # Initialize referenced sheet if needed
                                    if ref_sheet not in self.sheet_dependencies:
                                        self.sheet_dependencies[ref_sheet] = {
                                            "depends_on": set(),
                                            "referenced_by": set()
                                        }
                                    
                                    # Referenced sheet is referenced by current sheet
                                    self.sheet_dependencies[ref_sheet]["referenced_by"].add(sheet_name)
                    except Exception as e:
                        logger.warning(f"Error analyzing cell {row},{col} in sheet {sheet_name}: {str(e)}")
            
            logger.info(f"Sheet {sheet_name}: Found {formula_count} formulas, {cross_sheet_count} with cross-sheet references")
        
        # Convert sets to lists for JSON serialization
        for sheet_name in self.sheet_dependencies:
            self.sheet_dependencies[sheet_name]["depends_on"] = list(self.sheet_dependencies[sheet_name]["depends_on"])
            self.sheet_dependencies[sheet_name]["referenced_by"] = list(self.sheet_dependencies[sheet_name]["referenced_by"])
        
        total_cross_refs = len(self.cell_dependencies)
        logger.info(f"Completed cross-sheet formula analysis: Found {total_cross_refs} cells with cross-sheet references")
        
        return {
            "sheet_dependencies": self.sheet_dependencies,
            "cell_dependencies": self.cell_dependencies
        }
    
    def _get_cell_formula(self, cell) -> str:
        """
        Get formula from cell, handling different openpyxl versions.
        
        Args:
            cell: The openpyxl cell object
            
        Returns:
            Formula string or empty string if no formula
        """
        if hasattr(cell, 'formula') and cell.formula:
            return cell.formula
        elif hasattr(cell, '_value') and isinstance(cell._value, str) and cell._value.startswith('='):
            return cell._value
        elif isinstance(cell.value, str) and cell.value.startswith('='):
            return cell.value
        return ""
    
    def _extract_sheet_references(self, formula: str) -> Set[str]:
        """
        Extract sheet names referenced in a formula.
        
        Args:
            formula: Excel formula string
            
        Returns:
            Set of sheet names referenced in the formula
        """
        # Remove the leading = if present
        if formula.startswith('='):
            formula = formula[1:]
        
        # Pattern to match sheet references: Sheet!A1 or 'Sheet Name'!A1
        # This handles both quoted (potentially with spaces) and unquoted sheet names
        pattern = r"('([^']+)'|([^'!,() ]+))!"
        
        matches = re.findall(pattern, formula)
        sheet_refs = set()
        
        for match in matches:
            # match will be a tuple with groups
            if match[1]:  # Quoted sheet name with possible spaces
                sheet_refs.add(match[1])
            elif match[2]:  # Unquoted sheet name
                sheet_refs.add(match[2])
        
        return sheet_refs
    
    def get_sheet_dependency_graph(self):
        """
        Get a dependency graph for all sheets in the workbook.
        
        Returns:
            Dictionary with sheet dependencies
        """
        if not self.sheet_dependencies:
            self.analyze_all_cross_sheet_formulas()
        
        return self.sheet_dependencies
    
    def get_cells_referencing_sheet(self, sheet_name: str) -> Dict[str, Any]:
        """
        Get all cells that reference a specific sheet.
        
        Args:
            sheet_name: Name of the sheet
            
        Returns:
            Dictionary of cells that reference the specified sheet
        """
        if not self.cell_dependencies:
            self.analyze_all_cross_sheet_formulas()
        
        referencing_cells = {}
        
        for cell_ref, deps in self.cell_dependencies.items():
            if sheet_name in deps["references"]:
                referencing_cells[cell_ref] = deps
        
        return referencing_cells

    def get_update_impact_analysis(self, updated_cells: List[str]) -> Dict[str, Any]:
        """
        Analyze the impact of updating specific cells on cross-sheet formulas.
        
        Args:
            updated_cells: List of cell references that will be updated (format: "Sheet!A1")
            
        Returns:
            Dictionary with impact analysis
        """
        if not self.cell_dependencies:
            self.analyze_all_cross_sheet_formulas()
        
        # Group updated cells by sheet
        sheets_with_updates = {}
        for cell_ref in updated_cells:
            if "!" in cell_ref:
                sheet_name = cell_ref.split("!")[0]
                if sheet_name not in sheets_with_updates:
                    sheets_with_updates[sheet_name] = []
                sheets_with_updates[sheet_name].append(cell_ref)
        
        # Find all sheets that depend on updated sheets
        impacted_sheets = set()
        for sheet_name in sheets_with_updates:
            if sheet_name in self.sheet_dependencies:
                for dependent_sheet in self.sheet_dependencies[sheet_name]["referenced_by"]:
                    impacted_sheets.add(dependent_sheet)
        
        # Find specific cells that might be impacted
        potentially_impacted_cells = {}
        for cell_ref, deps in self.cell_dependencies.items():
            cell_sheet = cell_ref.split("!")[0]
            if cell_sheet in impacted_sheets:
                for ref_sheet in deps["references"]:
                    if ref_sheet in sheets_with_updates:
                        if cell_ref not in potentially_impacted_cells:
                            potentially_impacted_cells[cell_ref] = {
                                "formula": deps["formula"],
                                "depends_on_updated_sheets": []
                            }
                        potentially_impacted_cells[cell_ref]["depends_on_updated_sheets"].append(ref_sheet)
        
        return {
            "updated_sheets": list(sheets_with_updates.keys()),
            "impacted_sheets": list(impacted_sheets),
            "potentially_impacted_cells": potentially_impacted_cells,
            "total_impacted_cells": len(potentially_impacted_cells)
        }

def enhance_mapping_with_dependencies(
    workbook_data: Dict[str, Any],
    cross_sheet_dependencies: Dict[str, Any]
) -> Dict[str, Any]:
    """
    Enhance the workbook mapping with cross-sheet dependency information.
    
    Args:
        workbook_data: The extracted workbook structure information
        cross_sheet_dependencies: The cross-sheet dependency information
        
    Returns:
        Enhanced workbook mapping
    """
    # Deep copy the workbook data to avoid modifying the original
    enhanced_data = workbook_data.copy()
    
    # Add cross-sheet dependency information
    if "workbook_properties" not in enhanced_data:
        enhanced_data["workbook_properties"] = {}
    
    enhanced_data["workbook_properties"]["cross_sheet_dependencies"] = cross_sheet_dependencies["sheet_dependencies"]
    
    # Add dependency information to individual sheets
    for sheet_name, sheet_data in enhanced_data.get("sheets", {}).items():
        if sheet_name in cross_sheet_dependencies["sheet_dependencies"]:
            sheet_deps = cross_sheet_dependencies["sheet_dependencies"][sheet_name]
            
            # Add dependency data to sheet properties
            if "sheet_properties" not in sheet_data:
                sheet_data["sheet_properties"] = {}
                
            sheet_data["sheet_properties"]["depends_on_sheets"] = sheet_deps["depends_on"]
            sheet_data["sheet_properties"]["referenced_by_sheets"] = sheet_deps["referenced_by"]
    
    return enhanced_data

async def analyze_workbook_with_cross_sheet_formulas(
    file_content: bytes,
    workbook_mapping: Dict[str, Any],
    is_initial_upload: bool
) -> Dict[str, Any]:
    """
    Enhance workbook analysis with cross-sheet formula information.
    
    Args:
        file_content: Binary content of the Excel file
        workbook_mapping: The existing mapping of the workbook
        is_initial_upload: Whether this is an initial upload or update
        
    Returns:
        Enhanced workbook mapping with dependency information
    """
    try:
        # Analyze cross-sheet formulas
        formula_analyzer = CrossSheetFormulaAnalyzer(file_content)
        cross_sheet_info = formula_analyzer.analyze_all_cross_sheet_formulas()
        
        # Enhance the mapping with dependency information
        enhanced_mapping = enhance_mapping_with_dependencies(workbook_mapping, cross_sheet_info)
        
        # For initial uploads, also include formula cells with cross-sheet references
        if is_initial_upload:
            # Add important cross-sheet formula cells to key metrics
            for cell_ref, cell_info in cross_sheet_info["cell_dependencies"].items():
                if "!" in cell_ref:
                    sheet_name, cell_coord = cell_ref.split("!")
                    
                    if sheet_name in enhanced_mapping["sheets"]:
                        sheet_data = enhanced_mapping["sheets"][sheet_name]
                        
                        # Ensure key_metrics exists
                        if "key_metrics" not in sheet_data:
                            sheet_data["key_metrics"] = {}
                        
                        # Add cross-sheet formula as a key metric
                        metric_name = f"CrossSheetFormula_{cell_coord}"
                        sheet_data["key_metrics"][metric_name] = {
                            "formula": cell_info["formula"],
                            "references_sheets": list(cell_info["references"]),
                            "found_in_cells": [cell_coord]
                        }
        
        return enhanced_mapping
        
    except Exception as e:
        logger.error(f"Error enhancing mapping with cross-sheet formulas: {str(e)}")
        # Return original mapping if enhancement fails
        return workbook_mapping