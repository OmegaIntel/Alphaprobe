import copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Dict, Any, List, Tuple, Optional, Set
import logging

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger("excel_format_preservation")

class ExcelTemplatePreserver:
    """A class to handle all aspects of Excel template preservation."""
    
    @staticmethod
    def get_openpyxl_version():
        """Detect the version of openpyxl being used."""
        import openpyxl
        version = getattr(openpyxl, '__version__', 'unknown')
        logger.info(f"Using openpyxl version: {version}")
        return version
    
    def __init__(self, template_file_content: bytes):
        """
        Initialize with template file content.
        
        Args:
            template_file_content: The binary content of the template Excel file
        """
        # Log the openpyxl version being used
        self.openpyxl_version = self.get_openpyxl_version()
        self.template_content = template_file_content
        self.template_workbook = None
        self.template_metadata = None
        self.load_template()
    
    def load_template(self):
        """Load the template workbook with all properties preserved."""
        try:
            # Create a BytesIO object from the file content
            file_buffer = BytesIO(self.template_content)
            
            # Load the workbook with all options enabled
            # For openpyxl 3.1.5, data_only=True gets calculated values which might be needed
            self.template_workbook = load_workbook(file_buffer, data_only=False, keep_vba=True)
            
            # Also load a version with data_only=True to get calculated values if needed
            file_buffer.seek(0)
            self.template_workbook_data_only = load_workbook(file_buffer, data_only=True, keep_vba=True)
            
            # Extract and store metadata
            self.template_metadata = {
                "has_macros": self.template_workbook.vba_archive is not None,
                "sheet_names": self.template_workbook.sheetnames,
                "active_sheet": self.template_workbook.active.title,
                "named_ranges": self._extract_named_ranges(),
                "sheet_properties": self._extract_sheet_properties()
            }
            
            logger.info(f"Template loaded successfully with {len(self.template_metadata['sheet_names'])} sheets")
            logger.info(f"Template has macros: {self.template_metadata['has_macros']}")
            
        except Exception as e:
            logger.error(f"Error loading template: {str(e)}")
            raise Exception(f"Failed to load Excel template: {str(e)}")
    
    def _extract_named_ranges(self) -> List[Dict]:
        """Extract all named ranges from the workbook."""
        named_ranges = []
        
        try:
            # Check if defined_names is an attribute and is iterable
            if hasattr(self.template_workbook, 'defined_names') and self.template_workbook.defined_names:
                for name in self.template_workbook.defined_names:
                    try:
                        # Handle both object types and string types
                        if isinstance(name, str):
                            # If name is directly a string
                            named_ranges.append({
                                "name": name,
                                "value": None,
                                "scope": None
                            })
                        else:
                            # If name is an object with attributes
                            named_ranges.append({
                                "name": name.name if hasattr(name, "name") else str(name),
                                "value": name.value if hasattr(name, "value") else None,
                                "scope": name.scope if hasattr(name, "scope") else None
                            })
                    except Exception as e:
                        # Log warning but continue processing other named ranges
                        logger.warning(f"Error extracting named range: {str(e)}")
                        continue
        except Exception as e:
            logger.warning(f"Error accessing defined names: {str(e)}")
        
        return named_ranges
    
    def _extract_sheet_properties(self) -> Dict[str, Dict]:
        """Extract properties of all sheets."""
        sheet_properties = {}
        
        for sheet_name in self.template_workbook.sheetnames:
            try:
                sheet = self.template_workbook[sheet_name]
                
                # Get sheet visibility
                is_hidden = sheet.sheet_state != 'visible'
                
                # Get sheet protection status
                is_protected = sheet.protection.sheet
                
                # Get column widths and row heights
                col_dimensions = {}
                for col_idx, col_dim in sheet.column_dimensions.items():
                    if col_dim.width is not None:
                        col_dimensions[col_idx] = {
                            "width": col_dim.width,
                            "hidden": col_dim.hidden
                        }
                
                row_dimensions = {}
                for row_idx, row_dim in sheet.row_dimensions.items():
                    if row_dim.height is not None:
                        row_dimensions[row_idx] = {
                            "height": row_dim.height,
                            "hidden": row_dim.hidden
                        }
                
                # Count conditional formatting rules
                cf_count = len(sheet.conditional_formatting)
                
                # Count data validations
                dv_count = len(sheet.data_validations.dataValidation) if hasattr(sheet, 'data_validations') and sheet.data_validations else 0
                
                # Store properties
                sheet_properties[sheet_name] = {
                    "hidden": is_hidden,
                    "protected": is_protected,
                    "column_dimensions_count": len(col_dimensions),
                    "row_dimensions_count": len(row_dimensions),
                    "conditional_formatting_count": cf_count,
                    "data_validation_count": dv_count,
                    "freeze_panes": sheet.freeze_panes
                }
                
            except Exception as e:
                logger.warning(f"Error extracting properties for sheet {sheet_name}: {str(e)}")
                sheet_properties[sheet_name] = {"error": str(e)}
        
        return sheet_properties
    
    def extract_cell_styles(self, sheet_name: str, cell_ref: str) -> Dict[str, Any]:
        """
        Extract all style properties of a specific cell.
        
        Args:
            sheet_name: Name of the sheet
            cell_ref: Cell reference (e.g., "A1")
            
        Returns:
            Dictionary containing all cell style properties
        """
        try:
            if sheet_name not in self.template_workbook.sheetnames:
                raise Exception(f"Sheet {sheet_name} not found in template")
            
            sheet = self.template_workbook[sheet_name]
            cell = sheet[cell_ref]
            
            # Extract all style properties
            style = {
                "font": {
                    "name": cell.font.name if hasattr(cell.font, 'name') else None,
                    "size": cell.font.size if hasattr(cell.font, 'size') else None,
                    "bold": cell.font.bold if hasattr(cell.font, 'bold') else None,
                    "italic": cell.font.italic if hasattr(cell.font, 'italic') else None,
                    "underline": cell.font.underline if hasattr(cell.font, 'underline') else None,
                    "strike": cell.font.strike if hasattr(cell.font, 'strike') else None,
                    "color": cell.font.color.rgb if hasattr(cell.font, 'color') and cell.font.color else None
                },
                "fill": {
                    "fill_type": cell.fill.fill_type if hasattr(cell.fill, 'fill_type') else None,
                    "start_color": cell.fill.start_color.rgb if hasattr(cell.fill, 'start_color') and cell.fill.start_color else None,
                    "end_color": cell.fill.end_color.rgb if hasattr(cell.fill, 'end_color') and cell.fill.end_color else None
                },
                "border": {
                    "left": {
                        "style": cell.border.left.style if hasattr(cell.border, 'left') and cell.border.left else None,
                        "color": cell.border.left.color.rgb if hasattr(cell.border, 'left') and cell.border.left and hasattr(cell.border.left, 'color') and cell.border.left.color else None
                    },
                    "right": {
                        "style": cell.border.right.style if hasattr(cell.border, 'right') and cell.border.right else None,
                        "color": cell.border.right.color.rgb if hasattr(cell.border, 'right') and cell.border.right and hasattr(cell.border.right, 'color') and cell.border.right.color else None
                    },
                    "top": {
                        "style": cell.border.top.style if hasattr(cell.border, 'top') and cell.border.top else None,
                        "color": cell.border.top.color.rgb if hasattr(cell.border, 'top') and cell.border.top and hasattr(cell.border.top, 'color') and cell.border.top.color else None
                    },
                    "bottom": {
                        "style": cell.border.bottom.style if hasattr(cell.border, 'bottom') and cell.border.bottom else None,
                        "color": cell.border.bottom.color.rgb if hasattr(cell.border, 'bottom') and cell.border.bottom and hasattr(cell.border.bottom, 'color') and cell.border.bottom.color else None
                    }
                },
                "alignment": {
                    "horizontal": cell.alignment.horizontal if hasattr(cell.alignment, 'horizontal') else None,
                    "vertical": cell.alignment.vertical if hasattr(cell.alignment, 'vertical') else None,
                    "wrap_text": cell.alignment.wrap_text if hasattr(cell.alignment, 'wrap_text') else None,
                    "indent": cell.alignment.indent if hasattr(cell.alignment, 'indent') else None,
                    "text_rotation": cell.alignment.text_rotation if hasattr(cell.alignment, 'text_rotation') else None
                },
                "protection": {
                    "locked": cell.protection.locked if hasattr(cell.protection, 'locked') else None,
                    "hidden": cell.protection.hidden if hasattr(cell.protection, 'hidden') else None
                },
                "number_format": cell.number_format
            }
            
            return style
        
        except Exception as e:
            logger.error(f"Error extracting cell styles for {sheet_name}!{cell_ref}: {str(e)}")
            return {}
    
    def update_cell_value_preserving_styles(self, sheet_name: str, cell_ref: str, new_value: Any) -> bool:
        """
        Update a cell's value while completely preserving its styles, formulas, etc.
        
        Args:
            sheet_name: Name of the sheet
            cell_ref: Cell reference (e.g., "A1")
            new_value: New value to set
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if sheet_name not in self.template_workbook.sheetnames:
                raise Exception(f"Sheet {sheet_name} not found in template")
            
            sheet = self.template_workbook[sheet_name]
            cell = sheet[cell_ref]
            
            # Check if cell has a formula - handle different versions of openpyxl
            has_formula = None
            
            # Try different ways to detect formulas
            if hasattr(cell, 'formula'):
                has_formula = cell.formula
            elif hasattr(cell, '_value') and isinstance(cell._value, str) and cell._value.startswith('='):
                has_formula = cell._value
            elif isinstance(cell.value, str) and cell.value.startswith('='):
                has_formula = cell.value
                
            # If cell has a formula, we preserve it instead of setting the value
            if has_formula and isinstance(has_formula, str) and has_formula.strip():
                logger.info(f"Cell {sheet_name}!{cell_ref} has formula: {has_formula} - preserving formula")
                return False  # We don't update formula cells
            
            # Store original number format
            original_number_format = cell.number_format
            
            # Update the value
            cell.value = new_value
            
            # Ensure number format is preserved (Excel sometimes changes format on value update)
            cell.number_format = original_number_format
            
            logger.info(f"Updated cell {sheet_name}!{cell_ref} to value: {new_value} with number format: {original_number_format}")
            return True
            
        except Exception as e:
            logger.error(f"Error updating cell {sheet_name}!{cell_ref}: {str(e)}")
            return False
    
    def update_cells_with_mapping(self, cell_mapping: Dict[str, Dict[str, Any]], update_file_mapping: Dict) -> Tuple[bool, List[str]]:
        """
        Update cells in the template based on mapping information.
        
        Args:
            cell_mapping: Mapping between metrics and cells in the template
            update_file_mapping: The mapping JSON from the update file
            
        Returns:
            Tuple of (success_flag, list_of_updates_made)
        """
        updated_cells = []
        preserved_formulas = []
        
        try:
            # Process each sheet in the update file
            for sheet_name, sheet_data in update_file_mapping.get("sheets", {}).items():
                metrics = sheet_data.get("metrics", {})
                
                # For each metric in the update file
                for metric_name, metric_data in metrics.items():
                    # Check if we have a mapping for this metric
                    if metric_name in cell_mapping:
                        # Get the mapping info
                        map_info = cell_mapping[metric_name]
                        template_sheet_name = map_info["sheet"]
                        cell_refs = map_info["cells"]
                        
                        # Get the value from the update file
                        new_value = metric_data.get("sample_value")
                        
                        # Only proceed if we have a valid value
                        if new_value is not None:
                            for cell_ref in cell_refs:
                                # Try to update the cell while preserving its properties
                                update_success = self.update_cell_value_preserving_styles(
                                    template_sheet_name, cell_ref, new_value
                                )
                                
                                if update_success:
                                    updated_cells.append(f"{template_sheet_name}!{cell_ref} ({metric_name}): {new_value}")
                                else:
                                    # Cell likely had a formula that we preserved
                                    preserved_formulas.append(f"{template_sheet_name}!{cell_ref} ({metric_name}): Formula preserved")
            
            logger.info(f"Updated {len(updated_cells)} cells")
            logger.info(f"Preserved formulas in {len(preserved_formulas)} cells")
            
            return len(updated_cells) > 0, updated_cells + preserved_formulas
        
        except Exception as e:
            logger.error(f"Error updating cells with mapping: {str(e)}")
            return False, [f"Error: {str(e)}"]
    
    def save_workbook(self) -> bytes:
        """
        Save the workbook to bytes.
        
        Returns:
            Binary content of the updated Excel file
        """
        try:
            # Create BytesIO object
            output = BytesIO()
            
            # Save the workbook to the BytesIO object
            self.template_workbook.save(output)
            
            # Get the bytes content
            output.seek(0)
            content = output.getvalue()
            
            logger.info(f"Successfully saved workbook, size: {len(content)} bytes")
            
            return content
        
        except Exception as e:
            logger.error(f"Error saving workbook: {str(e)}")
            raise Exception(f"Failed to save Excel workbook: {str(e)}")
    
    def get_workbook_properties(self) -> Dict[str, Any]:
        """
        Get properties of the workbook.
        
        Returns:
            Dictionary containing workbook properties
        """
        return self.template_metadata
    
    def copy_worksheet_completely(self, source_sheet_name: str, target_sheet_name: str = None) -> str:
        """
        Create a complete copy of a worksheet with all properties preserved.
        
        Args:
            source_sheet_name: Name of the source sheet to copy
            target_sheet_name: Optional name for the new sheet (generated if not provided)
            
        Returns:
            Name of the new sheet
        """
        try:
            if source_sheet_name not in self.template_workbook.sheetnames:
                raise Exception(f"Source sheet {source_sheet_name} not found")
            
            source_sheet = self.template_workbook[source_sheet_name]
            
            # Generate target sheet name if not provided
            if not target_sheet_name:
                base_name = f"{source_sheet_name}_copy"
                target_sheet_name = base_name
                counter = 1
                
                # Ensure unique name
                while target_sheet_name in self.template_workbook.sheetnames:
                    target_sheet_name = f"{base_name}_{counter}"
                    counter += 1
            
            # Create a new sheet
            target_sheet = self.template_workbook.create_sheet(title=target_sheet_name)
            
            # Copy all cells content and formatting
            for row in source_sheet.rows:
                for cell in row:
                    # Copy cell value or formula
                    target_cell = target_sheet.cell(
                        row=cell.row, 
                        column=cell.column,
                        value=cell.value
                    )
                    
                    # Copy formula if present
                    if cell.formula:
                        target_cell.formula = cell.formula
                    
                    # Copy formatting
                    target_cell.font = copy.copy(cell.font)
                    target_cell.border = copy.copy(cell.border)
                    target_cell.fill = copy.copy(cell.fill)
                    target_cell.number_format = cell.number_format
                    target_cell.alignment = copy.copy(cell.alignment)
                    target_cell.protection = copy.copy(cell.protection)
            
            # Copy column dimensions
            for key, dimension in source_sheet.column_dimensions.items():
                target_sheet.column_dimensions[key].width = dimension.width
                target_sheet.column_dimensions[key].hidden = dimension.hidden
            
            # Copy row dimensions
            for key, dimension in source_sheet.row_dimensions.items():
                target_sheet.row_dimensions[key].height = dimension.height
                target_sheet.row_dimensions[key].hidden = dimension.hidden
            
            # Copy merged cells
            for merged_cell_range in source_sheet.merged_cells.ranges:
                target_sheet.merge_cells(str(merged_cell_range))
            
            # Copy conditional formatting
            for cf_range, rules in source_sheet.conditional_formatting.items():
                target_sheet.conditional_formatting[cf_range] = copy.deepcopy(rules)
            
            # Copy sheet protection
            target_sheet.protection = copy.copy(source_sheet.protection)
            
            # Copy print settings
            target_sheet.page_setup = copy.copy(source_sheet.page_setup)
            target_sheet.page_margins = copy.copy(source_sheet.page_margins)
            
            # Copy freeze panes
            target_sheet.freeze_panes = source_sheet.freeze_panes
            
            logger.info(f"Successfully copied sheet {source_sheet_name} to {target_sheet_name}")
            return target_sheet_name
            
        except Exception as e:
            logger.error(f"Error copying worksheet {source_sheet_name}: {str(e)}")
            raise Exception(f"Failed to copy worksheet: {str(e)}")
    
    @staticmethod
    def process_template_update(template_file_content: bytes, update_file_mapping: Dict, cell_mapping: Dict) -> Tuple[bytes, List[str]]:
        """
        Static method to process a template update.
        
        Args:
            template_file_content: Binary content of the template Excel file
            update_file_mapping: Mapping from the update file
            cell_mapping: Mapping between metrics and cells
            
        Returns:
            Tuple of (updated template binary content, list of updates made)
        """
        # Create a template preserver instance
        preserver = ExcelTemplatePreserver(template_file_content)
        
        # Update cells based on the mapping
        success, updates = preserver.update_cells_with_mapping(cell_mapping, update_file_mapping)
        
        if not success:
            raise Exception("No cells were updated in the template")
        
        # Save the updated workbook
        updated_content = preserver.save_workbook()
        
        return updated_content, updates