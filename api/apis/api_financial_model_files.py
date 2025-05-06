import uuid
import os
import json
from typing import Optional, List
from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime
import boto3
from openpyxl import load_workbook
from tempfile import NamedTemporaryFile
from io import BytesIO
import httpx

from db_models.financial_model import FinancialModel
from db_models.session import get_db
from api.api_user import get_current_user
from db_models.users import User as UserModel

FINANCIAL_MODEL_BUCKET = "financial-model-files"

s3 = boto3.client(
    "s3",
    aws_access_key_id=os.environ["S3_ACCESS_KEY"],
    aws_secret_access_key=os.environ["S3_SECRET_KEY"],
    region_name=os.environ["S3_REGION"]
)

router = APIRouter()

async def call_openai(prompt: str) -> str:
    """
    Call OpenAI API for AI model responses.
    Uses the gpt-3.5-turbo model, which is the most cost-effective option.
    """
    try:
        print(f"🔄 Calling OpenAI API with prompt size: {len(prompt)}")
        
        headers = {
            "Authorization": f"Bearer {os.environ['OPENAI_API_KEY']}",
            "Content-Type": "application/json"
        }
        
        # Using gpt-3.5-turbo-instruct which is cheaper than the chat models
        payload = {
            "model": "gpt-3.5-turbo-instruct",
            "prompt": prompt,
            "max_tokens": 1024,
            "temperature": 0.2,  # Lower temperature for more deterministic mapping
            "top_p": 0.9
        }
        
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.post(
                "https://api.openai.com/v1/completions",
                headers=headers,
                json=payload
            )
            
            if response.status_code != 200:
                print(f"⚠️ OpenAI API Error Response: {response.text}")
                raise Exception(f"OpenAI API error: {response.status_code} - {response.text}")
                
            result = response.json()
            generated_text = result["choices"][0]["text"].strip()
            
            print(f"✅ OpenAI response received successfully")
            return generated_text
            
    except Exception as e:
        print(f"🔥 Error calling OpenAI API: {str(e)}")
        print("⚠️ Using fallback mapping instead")
        return create_fallback_mapping(prompt)
    
def create_fallback_mapping(prompt: str) -> str:
    """Create a basic mapping when AI is unavailable"""
    # Extract headers from the prompt
    headers = []
    try:
        for line in prompt.split("\n"):
            if line.startswith("- "):
                header_part = line[2:].split(":")[0].strip()
                if header_part:
                    headers.append(header_part)
    except:
        pass
    
    # Very basic heuristic mapping - enhance as needed
    mapping = {}
    for header in headers:
        header_lower = header.lower()
        if any(term in header_lower for term in ["revenue", "sales", "turnover", "income"]):
            mapping["Revenue"] = header
        elif any(term in header_lower for term in ["ebitda", "profit", "earnings", "operating profit"]):
            mapping["EBITDA"] = header
        elif any(term in header_lower for term in ["net", "pat", "profit after", "income after"]):
            mapping["Net Income"] = header
        elif any(term in header_lower for term in ["cost", "cogs", "goods sold"]):
            mapping["Cost of Goods Sold"] = header
        elif any(term in header_lower for term in ["gross", "gp"]):
            mapping["Gross Profit"] = header
        elif any(term in header_lower for term in ["opex", "operating expense"]):
            mapping["Operating Expenses"] = header
    
    # Return as JSON string
    return json.dumps(mapping)

def build_mapping_prompt(headers: list[str], values: list[str] = []):
    sample_fields = [
        "Revenue", "Cost of Goods Sold", "Gross Profit", "Operating Expenses", "EBITDA",
        "Depreciation", "EBIT", "Interest Expense", "Profit Before Tax", "Tax", "Net Income"
    ]
    header_value_pairs = "\n".join(
        [f"- {h}: {v}" for h, v in zip(headers, values)]
    ) if values else "\n".join([f"- {h}" for h in headers])

    return f"""
You are a financial analyst assistant. A user has uploaded a financial spreadsheet.

Below are the columns extracted from the Excel file:

{header_value_pairs}

Map each of the columns to one of the following standardized financial metrics if they match in meaning:

{', '.join(sample_fields)}

Return only a JSON object like this:

{{
  "Revenue": "Turnover",
  "EBITDA": "Operating Profit",
  "Net Income": "PAT"
}}

Do not explain anything. Only return the JSON.
""".strip()


@router.post("/api/financial-models/upload")
async def upload_financial_model(
    company_name: str = Form(...),
    is_initial_upload: bool = Form(...),
    file: UploadFile = File(...),
    reference_model_id: Optional[str] = Form(None),
    note: Optional[str] = Form(None),
    current_user: UserModel = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        print("🔁 Starting upload process...")
        user_id = current_user.id
        print(f"👤 User ID: {user_id}")

        file_ext = os.path.splitext(file.filename)[-1].lower()
        if file_ext not in [".xlsx", ".xls"]:
            raise HTTPException(status_code=400, detail="Only Excel files are supported")

        print(f"📂 Uploading file: {file.filename} (ext: {file_ext})")

        # Read file content once and store it
        file_content = file.file.read()
        print(f"📄 File size: {len(file_content)} bytes")

        model_group_id = None
        template_file_path = None

        if not is_initial_upload:
            print("🔗 Referencing an existing model group")
            if not reference_model_id:
                raise HTTPException(status_code=400, detail="Reference model ID required for updates")

            reference_model = db.query(FinancialModel).filter(
                FinancialModel.id == reference_model_id,
                FinancialModel.is_initial_upload == True
            ).first()

            if not reference_model:
                raise HTTPException(status_code=404, detail="Reference model not found")

            print(f"📌 Reference model found: {reference_model.id}")
            model_group_id = reference_model.id

            with NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_template:
                print(f"📥 Downloading template from S3: {reference_model.original_file_s3}")
                s3.download_fileobj(FINANCIAL_MODEL_BUCKET, reference_model.original_file_s3, temp_template)
                template_file_path = temp_template.name
                print(f"📁 Template saved at: {template_file_path}")

        # Upload original to S3 - create fresh buffer
        s3_key_original = f"uploads/{user_id}/{str(uuid.uuid4())}_{file.filename}"
        upload_buffer = BytesIO(file_content)  # Fresh buffer for S3 upload
        print(f"🚀 Uploading original to S3: {s3_key_original}")
        s3.upload_fileobj(upload_buffer, FINANCIAL_MODEL_BUCKET, s3_key_original)

        s3_key_updated = None
        mapping_json = None

        if not is_initial_upload:
            print("🔍 Processing updated model file...")
            
            # Create fresh BytesIO object for workbook processing
            input_bytes_copy = BytesIO(file_content)
            
            try:
                wb_input = load_workbook(input_bytes_copy, data_only=True)
                input_sheet = wb_input.active

                headers = [cell.value for cell in next(input_sheet.iter_rows(min_row=1, max_row=1))]
                values = [cell.value for cell in next(input_sheet.iter_rows(min_row=2, max_row=2))]

                prompt = build_mapping_prompt(headers, values)
                try:
                    # Use OpenAI instead of Bedrock
                    raw_mapping = await call_openai(prompt)
                    mapping_json = json.loads(raw_mapping)
                except Exception as e:
                    print(f"🔥 AI mapping failed: {str(e)}")
                    raise HTTPException(status_code=500, detail=f"AI mapping failed: {str(e)}")

                wb = load_workbook(template_file_path)
                sheet = wb.active

                field_to_cell = {"Revenue": "B5", "EBITDA": "B6", "Net Income": "B7"}  # Example mapping
                for standard_field, source_col_name in mapping_json.items():
                    if standard_field in field_to_cell and source_col_name in headers:
                        col_idx = headers.index(source_col_name)
                        value = values[col_idx] if col_idx < len(values) else 0
                        sheet[field_to_cell[standard_field]] = value

                with NamedTemporaryFile(delete=False, suffix=".xlsx") as updated_file:
                    wb.save(updated_file.name)
                    updated_file.seek(0)
                    s3_key_updated = f"updated-models/{user_id}/{datetime.utcnow().isoformat()}_{file.filename}"
                    s3.upload_file(updated_file.name, FINANCIAL_MODEL_BUCKET, s3_key_updated)
                    print(f"📤 Updated model uploaded to: {s3_key_updated}")
            except Exception as e:
                print(f"🔥 Error processing updated model: {str(e)}")
                raise HTTPException(status_code=500, detail=f"Processing updated model failed: {str(e)}")

        # Store in DB
        new_model = FinancialModel(
            user_id=user_id,
            company_name=company_name,
            original_file_s3=s3_key_original,
            updated_model_s3=s3_key_updated,
            is_initial_upload=is_initial_upload,
            model_group_id=model_group_id,
            note=note,
            mapping_json=mapping_json
        )
        db.add(new_model)
        db.commit()
        db.refresh(new_model)

        if is_initial_upload:
            new_model.model_group_id = new_model.id
            db.add(new_model)
            db.commit()

        print("✅ Upload process completed.")
        return {
            "message": "File uploaded and processed",
            "data": {
                "id": str(new_model.id),
                "company_name": new_model.company_name,
                "original_file_s3": new_model.original_file_s3,
                "updated_model_s3": new_model.updated_model_s3,
                "model_group_id": str(new_model.model_group_id),
                "is_initial_upload": new_model.is_initial_upload,
                "mapping_json": new_model.mapping_json
            }
        }

    except Exception as e:
        print(f"🔥 Error during upload: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")
    
@router.get("/api/financial-models/group/{model_group_id}")
def get_models_by_group(model_group_id: str, db: Session = Depends(get_db), current_user: UserModel = Depends(get_current_user)):
    models = db.query(FinancialModel).filter(FinancialModel.model_group_id == model_group_id).order_by(FinancialModel.created_at.asc()).all()
    return {"models": [
        {
            "id": str(m.id),
            "company_name": m.company_name,
            "original_file_s3": m.original_file_s3,
            "updated_model_s3": m.updated_model_s3,
            "created_at": m.created_at.isoformat(),
            "is_initial_upload": m.is_initial_upload
        } for m in models
    ]}


@router.get("/api/financial-models/{id}")
def get_model(id: str, db: Session = Depends(get_db), current_user: UserModel = Depends(get_current_user)):
    model = db.query(FinancialModel).filter(FinancialModel.id == id).first()
    if not model:
        raise HTTPException(status_code=404, detail="Model not found")
    return {
        "id": str(model.id),
        "company_name": model.company_name,
        "original_file_s3": model.original_file_s3,
        "updated_model_s3": model.updated_model_s3,
        "is_initial_upload": model.is_initial_upload,
        "model_group_id": str(model.model_group_id) if model.model_group_id else None,
        "created_at": model.created_at.isoformat()
    }


@router.get("/api/financial-models/list")
def list_models(db: Session = Depends(get_db), current_user: UserModel = Depends(get_current_user)):
    models = db.query(FinancialModel).filter(FinancialModel.user_id == current_user.id).order_by(FinancialModel.created_at.desc()).all()
    return {"models": [
        {
            "id": str(m.id),
            "company_name": m.company_name,
            "original_file_s3": m.original_file_s3,
            "updated_model_s3": m.updated_model_s3,
            "created_at": m.created_at.isoformat(),
            "is_initial_upload": m.is_initial_upload,
            "model_group_id": str(m.model_group_id) if m.model_group_id else None
        } for m in models
    ]}


@router.delete("/api/financial-models/{id}")
def delete_model(id: str, db: Session = Depends(get_db), current_user: UserModel = Depends(get_current_user)):
    model = db.query(FinancialModel).filter(FinancialModel.id == id).first()
    if not model:
        raise HTTPException(status_code=404, detail="Model not found")
    db.delete(model)
    db.commit()
    return {"message": "Model deleted successfully"}